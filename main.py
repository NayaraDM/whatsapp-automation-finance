import os
import re
import time
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, numbers

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# ======== CONFIGURAÇÕES ========
# Nome do chat que será monitorado. Crie um chat/grupo no WhatsApp com este nome.
CHAT_NAME = os.getenv("WAPP_CHAT", "Minhas Compras")

# Caminho da planilha (será criada se não existir)
PLANILHA = Path("data") / "compras.xlsx"

# Intervalo entre varreduras (segundos)
POLL_SECONDS = 2.0
# ===============================


def inicializar_planilha():
    """Cria a planilha com cabeçalhos, se ainda não existir."""
    PLANILHA.parent.mkdir(parents=True, exist_ok=True)
    if not PLANILHA.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Compras"
        ws.append(["Data/Hora", "Item", "Valor", "Pagamento", "Observações", "Mensagem Original"])
        # Estilos simples
        for col in "ABCDEF":
            ws[f"{col}1"].font = Font(bold=True)
            ws[f"{col}1"].alignment = Alignment(horizontal="center")
        wb.save(PLANILHA)


def salvar_compra(item, valor_float, pagamento, observacoes, msg_original):
    """Adiciona uma nova linha na planilha."""
    try:
        wb = load_workbook(PLANILHA)
        ws = wb.active
    except FileNotFoundError:
        inicializar_planilha()
        wb = load_workbook(PLANILHA)
        ws = wb.active

    agora = datetime.now()
    linha = [
        agora.strftime("%d/%m/%Y %H:%M:%S"),
        item,
        valor_float,
        pagamento,
        observacoes,
        msg_original,
    ]
    ws.append(linha)

    # Formatar coluna Valor como moeda (R$)
    last_row = ws.max_row
    ws[f"C{last_row}"].number_format = numbers.FORMAT_CURRENCY_BRL_SIMPLE

    wb.save(PLANILHA)


def parse_mensagem(texto):
    """
    Aceita formatos como:
      - "Compra: Café; Valor: 12,50; Pgto: Pix; Obs: extra forte"
      - "Item: Arroz; Valor: 20; Pagamento: Cartão"
    Campos reconhecidos (case-insensitive): Compra/Item/Produto, Valor, Pgto/Forma/Pagamento, Obs/Observações
    Retorna dict com chaves: item, valor_float, pagamento, observacoes
    """
    original = texto.strip().replace("\n", " ")
    # Quebra por ';' para obter pares chave:valor
    partes = [p.strip() for p in original.split(";") if p.strip()]
    dados = {}

    for parte in partes:
        # separa na primeira ocorrência de ':'
        if ":" in parte:
            k, v = parte.split(":", 1)
            k = k.strip().lower()
            v = v.strip()

            if k in ("compra", "item", "produto"):
                dados["item"] = v
            elif k == "valor":
                # manter float independente de vírgula/ponto e símbolo de R$
                vf = v.replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
                try:
                    dados["valor_float"] = float(vf)
                except ValueError:
                    dados["valor_float"] = None
            elif k in ("pgto", "forma", "forma de pgto", "forma de pagamento", "pagamento"):
                dados["pagamento"] = v
            elif k in ("obs", "observacao", "observação", "observacoes", "observações"):
                dados["observacoes"] = v

    # Defaults
    item = dados.get("item") or ""
    valor_float = dados.get("valor_float")
    pagamento = dados.get("pagamento") or ""
    observacoes = dados.get("observacoes") or ""

    # Precisa ter pelo menos item e valor válidos para salvar
    if not item or valor_float is None:
        return None

    return {
        "item": item,
        "valor_float": valor_float,
        "pagamento": pagamento,
        "observacoes": observacoes,
        "msg_original": original,
    }


def abrir_whatsapp_e_ir_para_chat(driver, chat_name):
    driver.get("https://web.whatsapp.com")
    print("➡️  WhatsApp Web aberto. Escaneie o QR Code no celular (se aparecer).")

    # Aguarda a página principal (após login). Timeout generoso.
    wait = WebDriverWait(driver, 120)
    # Quando o topo/caixa de pesquisa estiver presente, consideramos logado
    wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    print("✅ Login concluído (ou sessão já estava ativa).")

    # Tenta achar e clicar o chat pelo título (nome do chat)
    try:
        print(f"➡️  Procurando chat '{chat_name}'…")
        chat_el = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located(
                (By.XPATH, f'//span[@title="{chat_name}"]')
            )
        )
        chat_el.click()
        print(f"✅ Chat '{chat_name}' aberto.")
    except Exception:
        print(f"⚠️  Não achei o chat '{chat_name}'.")
        print("   Abra manualmente o chat a ser monitorado e volte à janela do terminal.")
        input("   Pressione ENTER aqui quando o chat estiver aberto…")


def coletar_ultima_mensagem_recebida(driver):
    """
    Captura o texto da última mensagem recebida (da outra pessoa/conta),
    que contenha pelo menos 'Compra:' ou 'Item:' para filtrar ruído.
    """
    # Seleciona mensagens de entrada (message-in). Os seletores do WhatsApp podem mudar.
    mensagens = driver.find_elements(By.CSS_SELECTOR, "div.message-in span.selectable-text.copyable-text")
    if not mensagens:
        mensagens = driver.find_elements(By.CSS_SELECTOR, "div.message-in span.selectable-text")

    if not mensagens:
        return None

    # Pega a última mensagem que tenha nosso padrão
    for el in reversed(mensagens):
        txt = el.text.strip()
        if re.search(r"(?i)\b(compra|item|produto)\s*:", txt):
            return txt
    return None


def main():
    inicializar_planilha()

    # Inicia o Chrome (Selenium 4 tenta gerenciar o driver automaticamente)
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)

    try:
        abrir_whatsapp_e_ir_para_chat(driver, CHAT_NAME)

        print("\n🟢 Monitorando mensagens…")
        print("   Formato aceito: Compra: X; Valor: Y; Pgto: Z; Obs: ...\n")

        ultimo_processado = None

        while True:
            try:
                txt = coletar_ultima_mensagem_recebida(driver)
                if txt and txt != ultimo_processado:
                    dados = parse_mensagem(txt)
                    if dados:
                        salvar_compra(
                            dados["item"],
                            dados["valor_float"],
                            dados["pagamento"],
                            dados["observacoes"],
                            dados["msg_original"],
                        )
                        print(f"✅ Salvo: {dados['item']} | R$ {dados['valor_float']:.2f} | {dados['pagamento']} | {dados['observacoes']}")
                        ultimo_processado = txt
                    else:
                        # mensagem era 'COMPRA' mas faltou valor/item => ignorar
                        pass

                time.sleep(POLL_SECONDS)
            except KeyboardInterrupt:
                print("\n🛑 Encerrado pelo usuário.")
                break
            except Exception as e:
                print("Erro no loop:", e)
                time.sleep(2)

    finally:
        try:
            driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
